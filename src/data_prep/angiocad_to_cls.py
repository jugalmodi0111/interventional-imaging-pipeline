"""AngioCAD severity labels -> per-video classification records.

AngioCAD (Zenodo 10.5281/zenodo.15826856, CC-BY-4.0): 413 patients, ONE ROW EACH in
``AngioCAD_Labels.xlsx``, carrying a 7-grade stenosis severity for each of 15 coronary segments,
plus the video-series numbers for each side.

**It has NO bounding boxes** (verified 2026-08-23 by reading the 43 kB labels file — see
`docs/DATASETS.md`), so it cannot train the YOLO detector. It is a patient/segment-level
CLASSIFICATION dataset, and a good one: 74% of segment labels and 12.8% of patients are normal, so
real negatives exist at both granularities without having to mine them out of discarded videos the
way CADICA required (audit A1).

THE CORRECTNESS CONSTRAINT this module exists to enforce
--------------------------------------------------------
A video shows ONE coronary side. **83 of 413 patients (20%) have disease on one side and a wholly
normal other side** — 68 LCA-only, 15 RCA-only. Stamping the patient-level label onto every one of
their videos would label a healthy artery as diseased, teaching the model that a normal right
coronary is abnormal because the LAD happens to be occluded. That is the same class of error as
audit A1a (using lesion-video frames as negatives) and it would corrupt ~20% of the corpus.

So `video_label` restricts every label to the segments the video actually shows, resolved from the
sheet's own ``Right Coronary Series`` / ``Left Coronary Series`` columns. A series not listed in
either column yields ``None`` — refused, never guessed onto a side.

Segment-level and patient-level both fall out of the same record: `segments` is the 5- or 10-way
multi-label target for that view, and `positive` is `any(segments)` — the screening flag. Training
the segment head and deriving the flag keeps one model serving both granularities.

Pure stdlib + openpyxl (imported lazily inside `build_records`); no cv2/torch, so the label logic is
unit-testable with no heavy stack.
"""
import argparse
import json
import os

#: The 15 segments as the sheet names them, partitioned by which coronary a view shows.
RCA_SEGMENTS = ["Prox RCA", "Mid RCA", "Dist RCA", "PDA", "PLB"]
LCA_SEGMENTS = ["LM", "Prox LAD", "Mid LAD", "Dist LAD", "1st dig", "2nd dig",
                "Prox LCX", "Mid LCX", "Dist LCX", "OM"]

#: The seven severity grades, as CLOSED PERCENT BANDS. 'NL' is normal, not "unknown".
SEVERITY_BANDS = {
    "NL": (0, 0), "1-25": (1, 25), "26-50": (26, 50), "51-75": (51, 75),
    "76-90": (76, 90), "91-99": (91, 99), "100": (100, 100),
}


def parse_series_spec(spec):
    """``"1-2"`` / ``"7"`` / ``"3-9"`` -> the list of series numbers it names.

    openpyxl returns a bare number as an int, so ints are accepted too. Blank/None/unparseable
    yields ``[]`` rather than raising: a patient may simply have no series recorded for a side, and
    that is missing data, not a corrupt sheet.
    """
    if spec is None:
        return []
    if isinstance(spec, (int, float)):
        return [int(spec)]
    s = str(spec).strip()
    if not s:
        return []
    if "-" in s:
        lo, _, hi = s.partition("-")
        try:
            return list(range(int(lo.strip()), int(hi.strip()) + 1))
        except ValueError:
            return []
    try:
        return [int(s)]
    except ValueError:
        return []


def severity_band(grade):
    """Severity grade -> ``(low, high)`` percent band, or ``None`` if unrecognised.

    Returns None rather than ``(0, 0)`` for an unknown value on purpose: silently reading an
    unparseable grade as "normal" would turn a diseased artery into a training negative.
    """
    if grade is None:
        return None
    return SEVERITY_BANDS.get(str(grade).strip())


def is_significant(grade, threshold):
    """Is this segment diseased at or above ``threshold`` percent?

    Positive iff the ENTIRE band clears the threshold (``low >= threshold``). A grade that merely
    *might* reach it does not count -- ``'26-50'`` tops out at exactly 50 and so is not significant
    at a 50% cut. Using the band's low end keeps the positive class unambiguous; using the high end
    would make every borderline band positive and inflate prevalence.

    ``threshold`` is a CLINICAL choice (50% is common, 70% is the usual cut for intervention) and is
    not settled in this repo -- see `docs/STENOSIS_GATE_PROPOSAL.md`. Raises on an unknown grade so a
    sheet change surfaces loudly instead of quietly relabelling disease as normal.
    """
    band = severity_band(grade)
    if band is None:
        raise ValueError(f"unknown severity grade {grade!r}; expected one of {sorted(SEVERITY_BANDS)}")
    return band[0] >= threshold


def side_of_series(row, series):
    """``'rca'`` / ``'lca'`` for a video series number, or ``None`` if the sheet does not place it.

    Resolved from the row's own ``Right/Left Coronary Series`` columns. None means "this series is
    not listed for either side" -- the caller must drop the video rather than assign it a side,
    because a wrong side means labelling an artery with another artery's disease.
    """
    if series in parse_series_spec(row.get("Right Coronary Series")):
        return "rca"
    if series in parse_series_spec(row.get("Left Coronary Series")):
        return "lca"
    return None


def segments_for_side(side):
    """The segments visible in a view of ``side``."""
    return RCA_SEGMENTS if side == "rca" else LCA_SEGMENTS if side == "lca" else []


def video_label(row, series, threshold):
    """Label ONE video of ``row``'s patient -> ``{side, segments, positive}``, or None.

    ``segments`` maps only the segments THIS view shows to a bool at ``threshold``; ``positive`` is
    ``any(segments.values())`` -- the study-level screening flag derived from the segment head.
    Segments on the other coronary are deliberately absent, not False: they are unobserved in this
    view, and recording them as negatives would be as wrong as recording them as positives.

    Returns None when the series is not listed for either side (see `side_of_series`).
    """
    side = side_of_series(row, series)
    if side is None:
        return None
    segs = {}
    for seg in segments_for_side(side):
        grade = row.get(seg)
        if severity_band(grade) is None:      # blank/unknown cell -> unobserved, not normal
            continue
        segs[seg] = is_significant(grade, threshold)
    return {"side": side, "segments": segs, "positive": any(segs.values())}


def build_records(labels_path, threshold=50, frames_root=None):
    """Read ``AngioCAD_Labels.xlsx`` -> one record per (patient, video series).

    Each record carries ``patient``, ``series``, ``side``, ``segments``, ``positive`` and a
    ``group_key``. **``group_key`` is the PATIENT**, so every video of one patient lands on the same
    side of a train/val split -- the repo's standing rule (`io_utils.group_key`), and the reason the
    stenosis detector's honest F1 is 0.29 rather than the leaked 0.885.

    ``frames_root``, when given, is recorded per video as ``frames`` so a loader can find the PNGs;
    nothing here touches the image tree, so the label logic stays testable without the 16 GB.
    """
    import openpyxl                                            # lazy: keeps import cheap/torch-free

    wb = openpyxl.load_workbook(labels_path, read_only=True, data_only=True)
    ws = wb["Labels"] if "Labels" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]

    records = []
    for raw in rows:
        if raw is None or raw[0] is None:
            continue
        row = dict(zip(header, raw))
        patient = row.get("ID")
        for col in ("Right Coronary Series", "Left Coronary Series"):
            for series in parse_series_spec(row.get(col)):
                lab = video_label(row, series, threshold)
                if lab is None:
                    continue
                rec = {"patient": patient, "series": series, "group_key": f"angiocad_{patient}",
                       "threshold": threshold, **lab}
                if frames_root:
                    rec["frames"] = os.path.join(frames_root, str(patient), str(series))
                records.append(rec)
    return records


def main(labels_path, out_path, threshold=50, frames_root=None):
    recs = build_records(labels_path, threshold=threshold, frames_root=frames_root)
    with open(out_path, "w") as f:
        for r in recs:
            f.write(json.dumps(r, sort_keys=True) + "\n")
    pos = sum(r["positive"] for r in recs)
    pats = len({r["patient"] for r in recs})
    print(f"angiocad -> {out_path}: {len(recs)} videos / {pats} patients "
          f"({pos} positive, {len(recs) - pos} negative) at threshold {threshold}%")
    return recs


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="AngioCAD labels -> per-video classification records")
    ap.add_argument("--labels", required=True, help="AngioCAD_Labels.xlsx")
    ap.add_argument("--out", required=True, help="output JSONL")
    ap.add_argument("--threshold", type=int, default=50,
                    help="percent stenosis counted as significant (CLINICAL choice; 50 or 70)")
    ap.add_argument("--frames-root", default=None)
    a = ap.parse_args()
    main(a.labels, a.out, a.threshold, a.frames_root)
